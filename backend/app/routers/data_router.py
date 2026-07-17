import csv
import io
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import pandas as pd

from app.auth import get_current_user
from app.data_parsing import COLUMN_MAP, FileValidationError, parse_upload
from app.database import get_db
from app.ml.cache import clear_user
from app.ml.pipeline import MIN_RAW_ROWS, MIN_USABLE_ROWS, WARMUP_DAYS, build_features
from app.models import DemandRecord, User
from app.records import ALL_LOCATIONS, list_locations, load_records_df
from app.schemas import DataQualitySummary, DemandRecordOut, UploadResult

router = APIRouter(prefix="/api/data", tags=["data"])

SAMPLE_FILE = Path(__file__).resolve().parent.parent / "data" / "sample_template.xlsx"


def _save_dataframe(df, owner_id: int, db: Session) -> int:
    records = [
        DemandRecord(
            owner_id=owner_id,
            date=row.date,
            location=row.location if "location" in df.columns else None,
            demand=row.demand,
            avg_price=row.avg_price if "avg_price" in df.columns else None,
            cost_price=row.cost_price if "cost_price" in df.columns else None,
            production_volume=row.production_volume if "production_volume" in df.columns else None,
            season=row.season,
            is_holiday=bool(row.is_holiday),
            avg_temp=row.avg_temp if "avg_temp" in df.columns else None,
            rainfall=row.rainfall if "rainfall" in df.columns else None,
            tourists=row.tourists if "tourists" in df.columns else None,
            channel=row.channel,
            has_promotion=bool(row.has_promotion),
            note=row.note,
        )
        for row in df.itertuples(index=False)
    ]
    db.bulk_save_objects(records)
    db.commit()
    return len(records)


@router.post("/upload", response_model=UploadResult)
def upload_data(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Validate an uploaded file and, unless dry_run, replace the user's data.

    dry_run=true only parses and validates the file (no DB writes at all,
    existing data untouched) so the frontend can show a confirmation
    dialog with the exact impact before the user commits to replacing
    their data. dry_run=false performs the same validation and then
    replaces the data.
    """
    content = file.file.read()
    try:
        df, report = parse_upload(file.filename, content)
    except FileValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))

    existing_count = (
        db.query(DemandRecord).filter(DemandRecord.owner_id == current_user.id).count()
    )

    if dry_run:
        return UploadResult(dry_run=True, existing_rows_to_replace=existing_count, **report)

    db.query(DemandRecord).filter(DemandRecord.owner_id == current_user.id).delete()
    db.commit()

    imported = _save_dataframe(df, current_user.id, db)
    clear_user(current_user.id)
    return UploadResult(dry_run=False, existing_rows_to_replace=existing_count, **report)


@router.post("/load-sample", response_model=UploadResult)
def load_sample(
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    content = SAMPLE_FILE.read_bytes()
    try:
        df, report = parse_upload(SAMPLE_FILE.name, content)
    except FileValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))

    existing_count = (
        db.query(DemandRecord).filter(DemandRecord.owner_id == current_user.id).count()
    )

    db.query(DemandRecord).filter(DemandRecord.owner_id == current_user.id).delete()
    db.commit()

    imported = _save_dataframe(df, current_user.id, db)
    clear_user(current_user.id)
    return UploadResult(dry_run=False, existing_rows_to_replace=existing_count, **report)


@router.get("/records", response_model=list[DemandRecordOut])
def list_records(
    limit: int = Query(default=50, le=500),
    offset: int = 0,
    location: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(DemandRecord).filter(DemandRecord.owner_id == current_user.id)
    if location is not None and location != ALL_LOCATIONS:
        q = q.filter(DemandRecord.location == location)
    q = q.order_by(DemandRecord.date.desc()).offset(offset).limit(limit)
    return q.all()


@router.get("/locations")
def get_locations(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return {"locations": list_locations(db, current_user.id)}


@router.get("/summary")
def data_summary(
    location: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(DemandRecord).filter(DemandRecord.owner_id == current_user.id)
    if location is not None and location != ALL_LOCATIONS:
        q = q.filter(DemandRecord.location == location)
    count = q.count()
    if count == 0:
        return {"count": 0, "date_from": None, "date_to": None}
    first = q.order_by(DemandRecord.date.asc()).first()
    last = q.order_by(DemandRecord.date.desc()).first()
    return {"count": count, "date_from": first.date, "date_to": last.date}


@router.get("/quality", response_model=DataQualitySummary)
def data_quality(
    location: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    df = load_records_df(db, current_user.id, location=location)
    count = len(df)
    if count == 0:
        return DataQualitySummary(
            count=0,
            min_raw_rows_required=MIN_RAW_ROWS,
            ready_for_training=False,
            reason="ยังไม่มีข้อมูล กรุณาอัปโหลดไฟล์หรือโหลดข้อมูลตัวอย่างในหน้านี้ก่อน",
        )

    df["date"] = pd.to_datetime(df["date"])

    # Missing values among the exogenous columns that are optional in the
    # template but improve accuracy when present.
    optional_cols = ["avg_price", "cost_price", "production_volume", "avg_temp", "rainfall", "tourists"]
    missing_value_rows = int(df[optional_cols].isna().any(axis=1).sum())

    # With multi-location data the same date legitimately repeats once per
    # location, so only flag it as a duplicate within the same location.
    dedup_subset = ["date", "location"] if location is None and "location" in df.columns else ["date"]
    duplicate_date_rows = int(df.duplicated(subset=dedup_subset).sum())

    q1, q3 = df["demand"].quantile(0.25), df["demand"].quantile(0.75)
    iqr = q3 - q1
    lower_bound, upper_bound = q1 - 1.5 * iqr, q3 + 1.5 * iqr
    outlier_rows = int(((df["demand"] < lower_bound) | (df["demand"] > upper_bound)).sum())

    feat, _ = build_features(df)
    usable_rows = len(feat)
    ready = usable_rows >= MIN_USABLE_ROWS

    reason = None
    if not ready:
        reason = (
            f"มีข้อมูลดิบ {count} วัน แต่ต้องตัด {WARMUP_DAYS} วันแรกสำหรับสร้าง Lag/Rolling Mean "
            f"เหลือใช้ได้จริง {usable_rows} วัน ซึ่งน้อยกว่าขั้นต่ำที่ต้องการ ({MIN_USABLE_ROWS} วัน) "
            f"กรุณาเพิ่มข้อมูลอย่างน้อยรวม {MIN_RAW_ROWS} วัน"
        )

    return DataQualitySummary(
        count=count,
        date_from=df["date"].min().date(),
        date_to=df["date"].max().date(),
        missing_value_rows=missing_value_rows,
        duplicate_date_rows=duplicate_date_rows,
        outlier_rows=outlier_rows,
        usable_rows_for_training=usable_rows,
        min_raw_rows_required=MIN_RAW_ROWS,
        ready_for_training=ready,
        reason=reason,
    )


@router.delete("/records")
def clear_records(
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    deleted = (
        db.query(DemandRecord).filter(DemandRecord.owner_id == current_user.id).delete()
    )
    db.commit()
    clear_user(current_user.id)
    return {"deleted": deleted}


@router.get("/template")
def download_template():
    """Blank upload template: same columns as a real import, a couple of
    clearly-marked example rows showing the expected format for tricky
    fields (0/1 flags, season names, ...), and a short instructions sheet —
    for a first-time user who has never seen the column layout before."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ข้อมูล"

    headers = list(COLUMN_MAP.keys())
    ws.append(headers)
    header_fill = PatternFill(start_color="14664A", end_color="14664A", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    example_rows = [
        [
            "2026-01-01", 2000, 25, 20, 1800, "ฤดูหนาว", 0, 26, 5, 15000, "ค้าปลีก", "สมุทรสาคร", 0,
            "◄ แถวตัวอย่าง ลบก่อนนำเข้าข้อมูลจริง",
        ],
        [
            "2026-01-02", 2150, 25.5, 20, 1850, "ฤดูหนาว", 0, 26, 0, 15200, "ค้าส่ง", "สมุทรสาคร", 1,
            "◄ แถวตัวอย่าง ลบก่อนนำเข้าข้อมูลจริง",
        ],
    ]
    example_fill = PatternFill(start_color="FDF3E6", end_color="FDF3E6", fill_type="solid")
    for row in example_rows:
        ws.append(row)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = example_fill

    widths = [12, 20, 16, 20, 18, 12, 16, 14, 14, 16, 14, 12, 14, 28]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

    info = wb.create_sheet("คำแนะนำ")
    info_lines = [
        ("คำแนะนำการกรอกข้อมูล", True),
        ("", False),
        ("วันที่: รูปแบบ ปปปป-ดด-วว โดยใช้ปีคริสต์ศักราช (ค.ศ.) เช่น 2026-01-01 ไม่ใช่ปีพุทธศักราช (พ.ศ.) จำเป็นต้องกรอก", False),
        ("ความต้องการ/ยอดขาย (ลูก): ตัวเลข จำนวนลูกมะพร้าวที่ขายได้ในวันนั้น จำเป็นต้องกรอก", False),
        ("ราคาขายเฉลี่ย, ราคาหน้าสวน/ต้นทุน: ตัวเลข หน่วยบาทต่อลูก ไม่บังคับ", False),
        ("ปริมาณผลผลิต/สต๊อก: ตัวเลข จำนวนลูก ไม่บังคับ", False),
        ("ฤดูกาล: ระบุเป็นข้อความ เช่น ฤดูร้อน / ฤดูฝน / ฤดูหนาว ไม่บังคับ", False),
        ("วันหยุด/เทศกาล และ มีโปรโมชั่น: กรอก 0 (ไม่ใช่) หรือ 1 (ใช่) เท่านั้น ไม่บังคับ", False),
        ("อุณหภูมิเฉลี่ย, ปริมาณน้ำฝน, จำนวนนักท่องเที่ยว: ตัวเลข ไม่บังคับ", False),
        ("ช่องทางจำหน่าย: ระบุเป็นข้อความ เช่น ค้าส่ง / ค้าปลีก / ออนไลน์ ไม่บังคับ", False),
        ("จังหวัด: กรอกถ้าต้องการแยกพยากรณ์ตามพื้นที่ปลูก ปล่อยว่างได้ถ้ามีข้อมูลที่เดียว", False),
        ("หมายเหตุ: ข้อความอิสระ ไม่บังคับ", False),
        ("", False),
        ("ลบแถวตัวอย่าง (พื้นหลังสีส้มอ่อน) ในชีต \"ข้อมูล\" ก่อนกรอกข้อมูลจริงของคุณ แล้วนำไฟล์นี้ไปอัปโหลดที่หน้า \"ข้อมูล\" ในระบบ", False),
    ]
    for i, (text, bold) in enumerate(info_lines, start=1):
        cell = info.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=13, color="14664A")
    info.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=coconut_demand_template.xlsx"},
    )


@router.get("/export")
def export_csv(
    location: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(DemandRecord).filter(DemandRecord.owner_id == current_user.id)
    if location is not None and location != ALL_LOCATIONS:
        q = q.filter(DemandRecord.location == location)
    records = q.order_by(DemandRecord.date.asc()).all()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "date",
            "location",
            "demand",
            "avg_price",
            "cost_price",
            "production_volume",
            "season",
            "is_holiday",
            "avg_temp",
            "rainfall",
            "tourists",
            "channel",
            "has_promotion",
            "note",
        ]
    )
    for r in records:
        writer.writerow(
            [
                r.date,
                r.location,
                r.demand,
                r.avg_price,
                r.cost_price,
                r.production_volume,
                r.season,
                int(r.is_holiday),
                r.avg_temp,
                r.rainfall,
                r.tourists,
                r.channel,
                int(r.has_promotion),
                r.note,
            ]
        )
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=coconut_demand_export.csv"},
    )
